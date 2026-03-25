import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import shutil
import os


def process_gla(input_file, output_file="output_GLA.xlsx"):
    """
    Process GLA Excel file, write formatted output, return result dataframe.
    """

    # -------------------------------
    # Read file
    # -------------------------------
    df = pd.read_excel(input_file)

    # -------------------------------
    # Calculations
    # -------------------------------
    df["TAT for Payment Due date"] = (
        pd.to_datetime(df["NEXT_PAYDATE"], errors='coerce') -
        pd.to_datetime(df["PREMCESDTE"], errors='coerce')
    ).dt.days

    df["Protiviti GLA Calculation"] = np.where(
        (df["STATUS"] == "IF") & (df["CNTTYPE"].isin(["MSB", "SMB"])),
        (0.01 * df["ORIGINAL_SA"] * df["PREM_TERM"]).round(2),
        0.00
    )

    # -------------------------------
    # Copy original file to output
    # -------------------------------
    if hasattr(input_file, "read"):
        temp_input = "temp_input.xlsx"
        df.to_excel(temp_input, index=False)
        shutil.copy(temp_input, output_file)
        os.remove(temp_input)
    else:
        shutil.copy(input_file, output_file)

    # -------------------------------
    # Open workbook for formatting
    # -------------------------------
    wb = load_workbook(output_file)
    ws = wb.active

    AZ_COL  = 52
    BA_COL  = 53
    HEADER_ROW = 1

    header_style = Font(bold=True, name="Arial", size=10)
    header_fill  = PatternFill("solid", start_color="D9E1F2")

    # Write headers
    headers = [
        (AZ_COL, "TAT for Payment Due date"),
        (BA_COL, "Protiviti GLA Calculation")
    ]

    for col, label in headers:
        cell = ws.cell(row=HEADER_ROW, column=col, value=label)
        cell.font  = header_style
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Write data
    DATA_START_ROW = 2

    for i, (tat_val, gla_val) in enumerate(
        zip(df["TAT for Payment Due date"], df["Protiviti GLA Calculation"])
    ):
        r = DATA_START_ROW + i

        tat_cell = ws.cell(row=r, column=AZ_COL)
        tat_cell.value = int(tat_val) if pd.notna(tat_val) else None
        tat_cell.font  = Font(name="Arial", size=10)
        tat_cell.alignment = Alignment(horizontal="center")

        gla_cell = ws.cell(row=r, column=BA_COL)
        gla_cell.value = round(float(gla_val), 2) if pd.notna(gla_val) else None
        gla_cell.number_format = '#,##0.00'
        gla_cell.font  = Font(name="Arial", size=10)
        gla_cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["AZ"].width = 28
    ws.column_dimensions["BA"].width = 28

    wb.save(output_file)

    return df  # ← return for JSON preview
